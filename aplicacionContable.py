# -*- coding: utf-8 -*-
"""
Created on Wed Jan 12 21:36:59 2022

@author: carlo
"""

import tkinter as tk
import coneccionConBD as app
from tkinter import ttk
from tkcalendar import DateEntry
import openpyxl
from openpyxl.styles.borders import Border, Side

class ButtonAgregar:
    def __init__(self,buttonIngresar,rut,ventana,entradas,index):
        self.ingresar = buttonIngresar
        self.agregar = tk.Button(ventana, text = "+", command = lambda: self.Agregar(buttonIngresar,rut,ventana,entradas,index + 1))
        self.agregar.grid(row = index, column = 5)
        
    def Agregar(self,buttonIngresar,rut,ventana,entradas,index):
        self.agregar.destroy()
        self.agregar = None
        entrada = CrearEntradaTransaccion(rut,None,None,None,None,ventana,index)
        
        entradas.append(entrada)
        
        self.ingresar.grid(row = index + 1, column = 2)
        
        
        botonAgregar = ButtonAgregar(buttonIngresar,rut,ventana,entradas,index)
        
        botonBorrar = ButtonBorrar(buttonIngresar,botonAgregar,entrada,self.borrar,ventana,rut,entradas,index)
        
        botonAgregar.AsignarBorrar(botonBorrar)
        
        if self.borrar != None:
            self.borrar.AsignarSiguientes((entrada,botonAgregar,botonBorrar))
        
    def AsignarBorrar(self,buttonBorrar):
        self.borrar = buttonBorrar

class ButtonBorrar:
    def __init__(self,buttonIngresar,buttonAgregar,entrada,borrarAnterior,ventana,rut,entradas,index):
        self.ingresar = buttonIngresar
        self.agregar = buttonAgregar
        self.entrada = entrada
        self.borrarAnterior = borrarAnterior
        
        self.borrar = tk.Button(ventana, text = "-", command = lambda: self.Eliminar(buttonIngresar,rut,ventana,entradas,index - 1))
        self.borrar.grid(row = index, column = 6)
        self.entradasSiguientes = (None,None,None,None,None)
        
        
    def Eliminar(self,buttonIngresar,rut,ventana,entradas,index):
        
        entradas.remove(self.entrada)
        
        self.entrada.Destruir()
        
        botonAgregar = None
        
        if self.agregar != None and self.agregar.agregar != None:
            self.agregar.agregar.destroy()
            self.agregar = None
            
            botonAgregar = ButtonAgregar(buttonIngresar,rut,ventana,entradas,index)
            botonAgregar.AsignarBorrar(self.borrarAnterior)
            
        self.borrar.destroy()
        
        if self.borrarAnterior != None:
            if botonAgregar != None:
                self.borrarAnterior.AsignarAgregar(botonAgregar)
            else:
                self.borrarAnterior.AsignarAgregar(None)
            self.borrarAnterior.AsignarSiguientes(self.entradasSiguientes)
        
        
        
        indexIngresar = self.Reposicionar(index + 1)
        
        self.ingresar.grid(row = indexIngresar, column = 3)
        
        
    
    def AsignarSiguientes(self,entradaSiguiente):
        self.entradasSiguientes = entradaSiguiente
        
    def AsignarAgregar(self,buttonAgregar):
        self.Agregar = buttonAgregar
        
    def Reposicionar(self,index):
        
        if self.entradasSiguientes[0] != None:
            self.entradasSiguientes[0].AsignarIndex(index)
            if self.entradasSiguientes[1] != None and self.entradasSiguientes[1].agregar != None:    
                self.entradasSiguientes[1].agregar.grid(row = index, column = 5)
            
            self.entradasSiguientes[2].borrar.grid(row = index, column = 6)
            
            return self.entradasSiguientes[2].Reposicionar(index + 1)
        
        else:
            return index + 1

class EntradaTransacion:
    def __init__(self,debe,haber,cuenta):
        self.debe = debe
        self.haber = haber
        self.cuenta = cuenta
        
    def AsignarIndex(self,index):
        self.debe.grid(row = index, column = 2)
        self.haber.grid(row = index, column = 3)
        self.cuenta.grid(row = index, column = 1)
        
    def ObtenerDatos(self):
        return(int(self.debe.get()),int(self.haber.get()),self.cuenta.get().split()[0])
    
    def Destruir(self):
        self.debe.destroy()
        self.haber.destroy()
        self.cuenta.destroy()

def texto(top):
    
    data = app.MostrarCuentas('19.996.513-1')
    
    label = tk.Label(top, text = str(data))
    label.grid(row = 2, column = 0)

def recargar(data,combo,dictEmpresas):
    data = app.MostrarEmpresas()
    
    opciones = []
    
    dictEmpresas = {}
    
    for item in data:
        opciones.append(item[1] + " " + item[0])
        dictEmpresas[item[1]] = (item[0],item[1],item[2],item[3],item[4],item[5])
        
    combo["values"] = opciones
def recagarBorrarEmpresa(data,combo,dictEmpresas,comboGet):
    app.BorrarEmpresa(comboGet)
    
    recargar(data,combo,dictEmpresas)


def CrearPlantillaEmpresa(nombreText,duenyoText,rutText,rutPropText,direccionText,giroText,titulo):
    top = tk.Toplevel()
    top.title(titulo)
    
    nombreLabel = tk.Label(top, text = "nombre:")
    nombreLabel.grid(row = 0, column = 0)
    
    nombre = tk.Entry(top, width = 30)
    nombre.grid(row = 1, column = 0)
    if nombreText != None:
        nombre.insert(0,nombreText)
    
    duenyoLabel = tk.Label(top, text = "representante legal:")
    duenyoLabel.grid(row = 2, column = 0)
    
    duenyo = tk.Entry(top, width = 30)
    duenyo.grid(row = 3, column = 0)
    if duenyoText != None:
        duenyo.insert(0,duenyoText)
    
    rutLabel = tk.Label(top, text = "rut:")
    rutLabel.grid(row = 4, column = 0)
    
    rut = tk.Entry(top, width = 30)
    rut.grid(row = 5, column = 0)
    if rutText != None:
        rut.insert(0,rutText)
        
        
    rutPropLabel = tk.Label(top, text = "rut propietario:")
    rutPropLabel.grid(row = 6, column = 0)
    
    rutProp = tk.Entry(top, width = 30)
    rutProp.grid(row = 7, column = 0)
    if rutPropText != None:
        rutProp.insert(0,rutPropText)
    
    direccionLabel = tk.Label(top, text = "direccion:")
    direccionLabel.grid(row = 8, column = 0)
    
    direccion = tk.Entry(top, width = 30)
    direccion.grid(row = 9, column = 0)
    if direccionText != None:
        direccion.insert(0,direccionText)
        
    giroLabel = tk.Label(top, text = "giro:")
    giroLabel.grid(row = 10, column = 0)
    
    giro = tk.Entry(top, width = 30)
    giro.grid(row = 11, column = 0)
    if giroText != None:
        giro.insert(0,giroText)
    
    return top, nombre, duenyo, rut, rutProp, direccion, giro


def EditarEmpresa(datos):
    originalNombre, originalRut, originalDuenyo, originalRutProp,originalDireccion,originalGiro = datos
    
    if originalNombre == None:
        return
    top, nombre, duenyo, rut, rutProp, direccion, giro = CrearPlantillaEmpresa(originalNombre,originalDuenyo,originalRut, originalRutProp, originalDireccion, originalGiro,"editar empresa")
    
    btn = tk.Button(top, text = "ingresar", command = lambda: app.EditarEmpresa(nombre.get(),duenyo.get(),originalRut,rut.get(),rutProp.get(), direccion.get(), giro.get(),top))   
    btn.grid(row = 13, column = 0)


def recargarEmpresa(data,combo,dictEmpresas,nombre,duenyo,rut,rutProp,direccion,giro,top):
    app.CrearEmpresa(nombre,duenyo,rut,rutProp, direccion, giro,top)
    
    recargar(data,combo,dictEmpresas)

def crearEmpresa(data,combo,dictEmpresas):
    
    top, nombre, duenyo, rut, rutProp, direccion, giro = CrearPlantillaEmpresa(None,None,None,None,None,None,"crear empresa")
    
    btn = tk.Button(top, text = "ingresar", command = lambda: recargarEmpresa(data,combo,dictEmpresas,nombre.get(),duenyo.get(),rut.get(),rutProp.get(), direccion.get(), giro.get(),top))   
    btn.grid(row = 13, column = 0)
    
def CrearPlantillaCuenta(codigoText,nombreText,tipoText):
    top = tk.Toplevel()
    top.title("Añadir cuenta")
    
    nombreLabel = tk.Label(top, text = "nombre:")
    nombreLabel.grid(row = 0, column = 0,padx=10)
    
    nombre = tk.Entry(top, width = 30)
    nombre.grid(row = 0, column = 1,padx=5,pady=5)
    if nombreText != None:
        nombre.insert(0,nombreText)
    
    codigoLabel = tk.Label(top, text = "codigo:")
    codigoLabel.grid(row = 1, column = 0)
    
    codigo = tk.Entry(top, width = 30)
    codigo.grid(row = 1, column = 1,pady=5)
    if codigoText != None:
        codigo.insert(0,codigoText)
    
    tipoLabel = tk.Label(top, text = "tipo:")
    tipoLabel.grid(row = 2, column = 0)
    
    tipo = ttk.Combobox(top, state="readonly")
    tipo.grid(row = 2, column = 1,pady=5)
    
    tipo["values"] = ["activo","pasivo","patrimonio","ganacias","resultado"]
    
    if tipoText == "activo":
        tipo.current(0)
    elif tipoText == "pasivo":
        tipo.current(1)
    elif tipoText == "patrimonio":
        tipo.current(2)
    elif tipoText == "ganacias":
        tipo.current(3)
    elif tipoText == "resultado":
        tipo.current(4)
    
    return top, nombre, codigo, tipo    

def CrearCuenta(rut):
    top, nombre, codigo, tipo = CrearPlantillaCuenta(None,None,None)
    
    btn = tk.Button(top, text = "crear", command = lambda: app.CrearCuenta(rut,nombre.get(),codigo.get(),tipo.get(),top))   
    btn.grid(row = 3, column = 1,pady=10)
    
    
def SeleccionarCuenta(rut,operacion):
    
    top = tk.Toplevel()
    top.title("seleccionar cuenta")
    
    tipo = ttk.Combobox(top, state="readonly")
    tipo.grid(row = 1, column = 0, pady=5,padx=5)
    
    data = app.MostrarCuentas(rut)
    
    opciones = []
    
    datos = {}
    
    for item in data:
        opciones.append(item[1] + " " + item[0])
        datos[item[1]] = item
        
        
    tipo["values"] = opciones
    
    if operacion == "editar":
        btn = tk.Button(top, text = "seleccionar", command = lambda: EditarCuenta(rut,datos[tipo.get().split()[0]],top))   
        btn.grid(row = 2, column = 0, pady=5)
        
    elif operacion == "borrar":
        btn = tk.Button(top, text = "borrar", command = lambda: app.BorrarCuenta(rut,tipo.get().split()[0],top))   
        btn.grid(row = 2, column = 0, pady=5)
        
def EditarCuenta(rut,datos,ventana):
    ventana.destroy()
    
    nombreText, codigoText, tipoText = datos
    
    if nombreText == None:
        return
    
    top, nombre, codigo, tipo = CrearPlantillaCuenta(codigoText, nombreText, tipoText)
    
    btn = tk.Button(top, text = "editar", command = lambda: app.EditarCuenta(nombre.get(),codigo.get(),tipo.get(),rut,codigoText,top))   
    btn.grid(row = 2, column = 0)  
    
def VerCuentas(rut):
    top = tk.Toplevel()
    top.title("cuentas")
    
    datos = app.MostrarCuentas(rut)
    
    columns = ('#1', '#2', '#3')
    
    tree = ttk.Treeview(top, height = 10, columns=columns, show='headings')
    
    tree.heading('#1', text='codigo')
    tree.heading('#2', text='nombre')
    tree.heading('#3', text='tipo')
    
    tree.column("# 1", anchor = "center", width=100)
    tree.column("# 2", anchor = "center", width=100)
    tree.column("# 3", anchor = "center", width=100)
    tree.grid(row=0, column= 0, sticky='nsew')
    
    datosAux = []
    
    for item in datos:
        datosAux.append((item[0], item[1], item[2]))
        
        
    for d in datosAux:
        tree.insert('', tk.END, values=d)
        
    scrollbar = ttk.Scrollbar(top, orient=tk.VERTICAL, command=tree.yview)
    tree.configure(yscroll=scrollbar.set)
    scrollbar.grid(row=0, column=1, sticky='ns')


def CrearEntradaTransaccion(rut,cuentaCod,nombreCod,debeInt,haberInt,top,i):
    cuenta = ttk.Combobox(top, state="readonly")
    cuenta.grid(row = i, column = 1)
    #cuenta.place(x=10, y=10)
    
    data = app.MostrarCuentas(rut)
    
    opciones = []
    
    for item in data:
        opciones.append(item[1] + " " + item[0])
        
    cuenta["values"] = opciones
    
    if cuentaCod != None:
       ind = opciones.index(str(cuentaCod) + " " + str(nombreCod))
       cuenta.current(ind)
    
    debe = tk.Entry(top, width = 30)
    debe.grid(row = i, column = 2)
    if debeInt != None:
        debe.insert(0,debeInt)
    
    haber = tk.Entry(top, width = 30)
    haber.grid(row = i, column = 3)
    
    if haberInt != None:
        haber.insert(0,haberInt)
        
    entrada = EntradaTransacion(debe, haber, cuenta)
    
    return entrada

def des(top):
    top.destroy()

def MensajeWindow(mensaje):
    top = tk.Toplevel()
    top.title("mensaje")
    
    texto = tk.Label(top, text = mensaje)
    texto.grid(row = 1, column = 1)
    
    ok = tk.Button(top, text = "OK", command = lambda: des(top))
    ok.grid(row = 2, column = 1)

def IngresarComprobante(rut,fecha,tipo,glosa,entradas,ventana):
    
    if app.IngresarComprobante(rut,fecha,tipo,glosa,entradas,ventana):
        MensajeWindow("comprobante ingresado")
    else:
        MensajeWindow("su asiento no cuadra")
        
def ObtenerDatosTransacciones(rut,codigoCom):

    db = app.Conectar()
    cursor = db.cursor()
    
    sql = "SELECT t.debe,t.haber,t.codigo,nombre FROM transaccion t inner join cuenta c on t.codigo = c.codigo and t.rut = c.rut WHERE c.rut = '{0}' and codigoCom = {1};".format(rut,codigoCom)
    
    resultsTransaccion = None
    try:
        # Execute the SQL command
        cursor.execute(sql)
        
        resultsTransaccion = cursor.fetchall()
    except:
        print("error en la base de datos")
        db.rollback()
        app.Desconectar(db)
        return None,None,None  
    
    return resultsTransaccion
        
def EditarComprobante(rut,datos):#pendiente

    codigoCom, glosa, tipo, fecha = datos
    
    transaccionData = ObtenerDatosTransacciones(rut,codigoCom)
    
    fecha,tipo,glosa,entradas,top,_ = PlantillaComprobante(rut,fecha,tipo,glosa,transaccionData,"editar comprobante",codigoCom,False)

        
def EditarComprobante2(rut,codigoCom,fecha,tipo,glosa,entradas,top):
    if app.EditarComprobante(rut,codigoCom,fecha,tipo,glosa,entradas,top):
        MensajeWindow("comprobante ingresado")
    else:
        MensajeWindow("su asiento no cuadra")

def PlantillaComprobante(rut,fechaText,tipoText,glosaText,transaccionesText,titulo,codigoCom,ingresarOperation):
    top = tk.Toplevel()
    top.title(titulo)
    
    fechaLabel = tk.Label(top, text = "fecha:")
    fechaLabel.grid(row = 1, column = 1)
    
    fecha = DateEntry(top, width= 16, background= "magenta3", foreground= "white",bd=2)
    fecha.grid(row = 1, column = 2)
    fecha.set_date(fechaText)
    
    tipoLabel = tk.Label(top, text = "tipo:")
    tipoLabel.grid(row = 2, column = 1)
    
    tipo = ttk.Combobox(top, state="readonly")
    tipo.grid(row = 2, column = 2)
    #tipo.place(x=100, y=50)
    
    tipo["values"] = ["E","I","T"]
    
    if tipoText == "E":
        tipo.current(0)
    elif tipoText == "I":
        tipo.current(1)
    elif tipoText == "T":
        tipo.current(2)
    
    glosaLabel = tk.Label(top, text = "glosa:")
    glosaLabel.grid(row = 3, column = 1)
    
    glosa = tk.Entry(top, width = 30)
    glosa.grid(row = 3, column = 2)
    
    if glosaText != None:
        glosa.insert(0,glosaText)
        
    cuentaLabel = tk.Label(top, text = "cuenta")
    cuentaLabel.grid(row = 5, column = 1)
    
    debeLabel = tk.Label(top, text = "debe")
    debeLabel.grid(row = 5, column = 2)
    
    haberLabel = tk.Label(top, text = "haber")
    haberLabel.grid(row = 5, column = 3)
    
    entradas = []
   # borrarButtons = []
    
    i = 6
    ingresar = None
    
   
    if ingresarOperation:
        ingresar = tk.Button(top, text = "ingresar", command = lambda: IngresarComprobante(rut,fecha.get_date(),tipo.get(),glosa.get(),entradas,top))
    else:
        ingresar = tk.Button(top, text = "ingresar", command = lambda: EditarComprobante2(rut,codigoCom,fecha.get_date(),tipo.get(),glosa.get(),entradas,top))
    
    
    borrarAnterior = None
    count = 0
    for item in transaccionesText:
        count += 1
        aux = CrearEntradaTransaccion(rut,item[2],item[3],item[0],item[1],top,i)
        entradas.append(aux)
        
        botonBorrar = ButtonBorrar(ingresar,None,aux,borrarAnterior,top,rut,entradas,i)
        
        if borrarAnterior != None:
            borrarAnterior.AsignarSiguientes((aux, None,botonBorrar))
        
        borrarAnterior = botonBorrar

        i += 1
    
    if ingresarOperation:
        aux = CrearEntradaTransaccion(rut,None,None,None,None,top,i)
        entradas.append(aux)
        i += 1
        
        
    botonAgregar = ButtonAgregar(ingresar,rut,top,entradas,i - 1)
    botonAgregar.AsignarBorrar(None)
    
    i += 1
    ingresar.grid(row = i - 1, column = 2)
    
    return fecha,tipo,glosa,entradas,top,i + 1
    

def CrearComprobante(rut):
    fecha,tipo,glosa,entradas,top, index = PlantillaComprobante(rut,None,None,None,[],"crear comprobante",None,True)

def SeleccionarComprobante(rut,operacion):
    top = tk.Toplevel()
    top.title("seleccionar comprobante")
     
    tipo = ttk.Combobox(top, state="readonly")
    tipo.grid(row = 0, column = 0, padx=5, pady=5)

    com = app.ObtenerComprobantes(rut)
     
    opciones = []
    
    data = {}
     
    for item in com:
        opciones.append(item[1])
        data[item[1]] = (item[0],item[1],item[2],item[3])
        
    tipo["values"] = opciones
    
    if operacion == "editar":
        btn = tk.Button(top, text = "seleccionar", command = lambda: EditarComprobante(rut,data[tipo.get()]))   
        btn.grid(row = 1, column = 0, pady=5)
        
    elif operacion == "borrar":
        btn = tk.Button(top, text = "borrar", command = lambda: recargarBorrarComrprobante(rut,data,tipo))   
        btn.grid(row = 1, column = 0, pady=5)
        
    elif operacion == "ver":
        btn = tk.Button(top, text = "seleccionar", command = lambda: VerComprobante(rut, data[tipo.get()]))   
        btn.grid(row = 1, column = 0, pady=5)

def VerComprobante(rut,datos):
    top = tk.Toplevel()
    top.title("ver comprobante")

    dataTransaccion = app.ObtenerTransacciones(rut, datos[0])
    
    fechaLabel = tk.Label(top, text = "fecha: " + str(datos[3]))
    fechaLabel.grid(row = 1, column = 1)
    
    glosaLabel = tk.Label(top, text = "glosa: " + str(datos[1]))
    glosaLabel.grid(row = 1, column = 2)
    
    tipoLabel = tk.Label(top, text = "tipo: " + str(datos[2]))
    tipoLabel.grid(row = 1, column = 3)
    
    i = 2
    
    for item in dataTransaccion:
        debeLabel = tk.Label(top, text = "debe: " + str(item[0]))
        debeLabel.grid(row = i, column = 1)
        
        haberLabel = tk.Label(top, text = "haber: " + str(item[1]))
        haberLabel.grid(row = i, column = 2)
        
        codigoLabel = tk.Label(top, text = "codigo cuenta: " + str(item[2]))
        codigoLabel.grid(row = i, column = 3)

        nombreLabel = tk.Label(top, text = "nombre cuenta: " + str(item[3]))
        nombreLabel.grid(row = i, column = 4)
    
        tipoLabel = tk.Label(top, text = "tipo cuenta: " + str(item[4]))
        tipoLabel.grid(row = i, column = 5)
    
        i += 1

def recargarBorrarComrprobante(rut,data,tipo):
    app.BorrarComprobante(rut, data[tipo.get()][0])
    
    recargarComprobante(data,rut,tipo)

def recargarComprobante(data,rut,tipo):
    com = app.ObtenerComprobantes(rut)
    opciones = []
    data = {}
     
    for item in com:
        opciones.append(item[1])
        data[item[1]] = (item[0],item[1],item[2],item[3])
        
    tipo["values"] = opciones
    
    
def MostrarTransacciones(transacciones,top,isMayor,rut,fechaInicio,fechaFinal):
    
    datosExcel = {}
    
    datosFecha = []
    datosComprobante = []
    datosGlosa = []
    datosDebe = []
    datosHaber = []
    datosCodigoCuenta = []
    datosNombreCuenta = []
    
    columns = ('#1', '#2', '#3','#4','#5','#6','#7')
    
    tree = ttk.Treeview(top, height = 10, columns=columns, show='headings')
    
    tree.heading('#1', text='fecha')
    tree.heading('#2', text='comprobante')
    tree.heading('#3', text='glosa')
    tree.heading('#4', text='debe')
    tree.heading('#5', text='haber')
    tree.heading('#6', text='codigo')
    tree.heading('#7', text='nombre cuenta')
    
    tree.column("# 1", anchor = "center", width=100)
    tree.column("# 2", anchor = "center", width=100)
    tree.column("# 3", anchor = "center", width=100)
    tree.column("# 4", anchor = "center", width=100)
    tree.column("# 5", anchor = "center", width=100)
    tree.column("# 6", anchor = "center", width=100)
    tree.column("# 7", anchor = "center", width=100)
    
    tree.grid(row=0, columnspan= 30, sticky='nsew')
    
    datosExcelTotales = {}
    
    datosAux = []
    
    totalDebe = 0
    totalHaber = 0
    
    for item in transacciones:
        datosAux.append((item[0], item[1], item[2], item[3], item[4], item[5], item[6]))
        datosFecha.append(str(item[0]))
        datosComprobante.append(str(item[1]))
        datosGlosa.append(str(item[2]))
        datosDebe.append(str(item[3]))
        datosHaber.append(str(item[4]))
        datosCodigoCuenta.append(str(item[5]))
        datosNombreCuenta.append(str(item[6]))
        totalDebe += item[3]   
        totalHaber += item[4]
        
    for d in datosAux:
        tree.insert('', tk.END, values=d)
        
    scrollbar = ttk.Scrollbar(top, orient=tk.VERTICAL, command=tree.yview)
    tree.configure(yscroll=scrollbar.set)
    scrollbar.grid(row=0, column=31, sticky='ns')
       
        
    datosExcel["fecha"] = datosFecha
    datosExcel["comprobante"] = datosComprobante
    datosExcel["glosa"] = datosGlosa
    datosExcel["debe"] = datosDebe
    datosExcel["haber"] = datosHaber
    datosExcel["codigo cuenta"] = datosCodigoCuenta
    datosExcel["nombre cuenta"] = datosNombreCuenta
    
    datosExcelTotales["total debe"] = totalDebe
    datosExcelTotales["total haber"] = totalHaber
    
    nombreNadaTotal = tk.Label(top, text = " ")
    nombreNadaTotal.grid(row = 1, column = 0)
    
    nombreTextTotalDebe = tk.Label(top, text = "total debe: ")
    nombreTextTotalDebe.grid(row = 2, column = 28)
  
    nombreTotalDebe = tk.Label(top, text = str(totalDebe))
    nombreTotalDebe.grid(row = 2, column = 29)
    
    nombreTextTotalHaber = tk.Label(top, text = "total haber: ")
    nombreTextTotalHaber.grid(row = 3, column = 28)
    
    nombreTotalHaber = tk.Label(top, text = str(totalHaber))
    nombreTotalHaber.grid(row = 3, column = 29)
    
    generar = tk.Button(top, text = "generar excel", command = lambda: GenerarExcelDiario(datosExcel,datosExcelTotales,rut,fechaInicio,fechaFinal))
    generar.grid(row = 2, column = 0)
  
def PrepararExcel(rut,nombreHoja):
    
    wb = openpyxl.Workbook()
    
    hoja = wb.active
    hoja.title = nombreHoja
    
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    
    up_thin_border = Border(top=Side(style='thin'))
    
    return wb, hoja, thin_border, up_thin_border    

def GenerarExcelDiario(datos,datosTotales,rut,fechaInicio,fechaFinal):
    
    
    wb, hoja, thin_border, up_thin_border = PrepararExcel(rut,'diario')
    
    sheet = wb.get_sheet_by_name('diario')
    
    fechaInicioAux = fechaInicio.split('/')
    fechaFinalAux = fechaFinal.split('/')
    
    sheet['B2'].value = "DIARIO GENERAL"
    sheet['B3'].value = "periodo: entre " + fechaInicioAux[1] + "/" + fechaInicioAux[0] + "/" + fechaInicioAux[2] + " y " + fechaFinalAux[1] + "/" + fechaFinalAux[0] + "/" + fechaFinalAux[2]
    sheet.merge_cells(start_row=2, start_column=2, end_row=2, end_column=9)
    
    
    sheet['B5'].value = "fecha"
    sheet['C5'].value = "comprobante"
    sheet['D5'].value = "glosa"
    sheet['E5'].value = "debe"
    sheet['F5'].value = "haber"
    sheet['G5'].value = "codigo cuenta"
    sheet['H5'].value = "cuenta"
    
    sheet['B6'].border = up_thin_border
    sheet['C6'].border = up_thin_border
    sheet['D6'].border = up_thin_border
    sheet['E6'].border = up_thin_border
    sheet['F6'].border = up_thin_border
    sheet['G6'].border = up_thin_border
    sheet['H6'].border = up_thin_border
    

    columnas = ['B','C','D','E','F','G','H']
    j = 0
    for key in datos:
        i = 6
        for d in datos[key]:
            sheet[columnas[j] + str(i)].value = d
            i += 1
            
        j += 1
        
    sheet['C' + str(i)].value = "TOTAL"
    
    j = 3
    
    for key in datosTotales:
        sheet[columnas[j] + str(i)].value = int(datosTotales[key])
        
        j += 1
        
    for c in columnas:
        sheet[c + str(i)].border = up_thin_border
        
    
    nombre = str(rut) + " " + str(fechaInicio) + " " + str(fechaFinal) + " diario.xlsx"
    nombre = nombre.replace("/", "-")
    
    wb.save(nombre)
    
    print("excel creado")    

def VerDiario(rut,fechaInicio,fechaFinal):
    top = tk.Toplevel()
    top.title("Diario")
    datos = app.ObtenerDiario(rut,fechaInicio,fechaFinal)
    
    MostrarTransacciones(datos,top,False,rut,fechaInicio,fechaFinal)
    
def SeleccionarFecha(rut,operacion,frames):
    top = tk.Toplevel()
    top.title(operacion)
    
    fechaInicioLabel = tk.Label(top, text = "fecha de inicio:")
    fechaInicioLabel.grid(row = 1, column = 1)
    
    fechaInicio = DateEntry(top, width= 16, background= "magenta3", foreground= "white",bd=2)
    fechaInicio.grid(row = 1, column = 2)
    
    fechaFinalLabel = tk.Label(top, text = "fecha final:")
    fechaFinalLabel.grid(row = 2, column = 1)
    
    fechaFinal = DateEntry(top, width= 16, background= "magenta3", foreground= "white",bd=2)
    fechaFinal.grid(row = 2, column = 2)
    
    if operacion == "diario":
    
        ingresarButton = tk.Button(top,text = "ingresar", command = lambda: VerDiario(rut,fechaInicio.get(),fechaFinal.get()))
    
    elif operacion == "mayor":
        tipo = ttk.Combobox(top, state="readonly")
        tipo.place(x=5, y=50)
        data = app.MostrarCuentas(rut)
    
        opciones = []
    
        datos = {}
    
        for item in data:
            opciones.append(item[1] + " " + item[0])
            datos[item[1]] = item
        
        opciones.append("todos")
        
        tipo["values"] = opciones
        
        ingresarButton = tk.Button(top,text = "ingresar", command = lambda: proxiMayor(rut,fechaInicio.get(),fechaFinal.get(),datos,tipo,frames))
        
    elif operacion == "balance":
        ingresarButton = tk.Button(top,text = "ingresar", command = lambda: VerBalance(rut,fechaInicio.get(),fechaFinal.get()))
    
    ingresarButton.grid(row = 3,column = 3, padx=5, pady=5)
    
def obtenerCodigoMayor(elemen):
    return elemen[5]
    
def VerMayor(rut,fechaInicio,fechaFinal,codCuenta,frames):
    top = tk.Toplevel()
    top.title("Mayor")
    
    datos2 = app.ObtenerMayor(rut,codCuenta,fechaInicio,fechaFinal)
    
    datos = []
    
    for d in datos2:
        datos.append(d)
        
    datos.sort(key=obtenerCodigoMayor)
    
    listaVentanasCuentas = []
    
    cuentasDeMayor = {}
    cuentasDeMayorTotales = {}
    
    global frameMostrado
    
    frameMostrado = tk.LabelFrame(top, text = "inicio", padx = 10, pady = 10)
    
    for item in datos:
        if item[6] not in cuentasDeMayor:
            cuentasDeMayor[item[6]] = []
            auxFr = tk.LabelFrame(top, text = str(item[7]), padx = 10, pady = 10)
            listaVentanasCuentas.append(auxFr)
            
        cuentasDeMayor[item[6]].append(item)
           
    
    
    i = 0
    
    saldosCuentas = {}
    
    for c in cuentasDeMayor:
        auxList = cuentasDeMayor[c]
        cuentasDeMayorTotales[c] = []
        saldosCuentas[c] = []
        mostrarMayorCuenta2(auxList, auxList[0][6],top, listaVentanasCuentas[i],cuentasDeMayorTotales[c],saldosCuentas[c])
        i += 1
    
    
    
    for fr in listaVentanasCuentas:
        siguiente = tk.Button(fr, text = "siguiente", command = lambda: siguienteFrame(listaVentanasCuentas))
        siguiente.grid(row = 2, column = 0, sticky='NW')
        
        anterior = tk.Button(fr, text = "anterior", command = lambda: anteriorFrame(listaVentanasCuentas))
        anterior.grid(row = 2, column = 1, sticky='NW')
        
        crearExcel = tk.Button(fr, text = "crear excel", command = lambda: GenerarExcelMayor(cuentasDeMayor,cuentasDeMayorTotales,saldosCuentas,rut,fechaInicio,fechaFinal))
        crearExcel.grid(row = 2, column = 2, sticky='NW')
    
    
    frameMostrado.grid_forget()
    frameMostrado = listaVentanasCuentas[0]
    frameMostrado.grid(row = 0, column = 0)
    
indexFrame = 0
frameMostrado = None
    
def anteriorFrame(lista):
    global indexFrame
    indexFrame -= 1
    if indexFrame < 0:
        indexFrame = len(lista) - 1
    
    global frameMostrado
    
    frameMostrado.grid_forget()
    frameMostrado = lista[indexFrame]
    frameMostrado.grid(row = 0, column = 0)
    
def siguienteFrame(lista):
    global indexFrame
    indexFrame += 1
    if indexFrame >= len(lista):
        indexFrame = 0
    
    global frameMostrado    
    
    frameMostrado.grid_forget()
    frameMostrado = lista[indexFrame]
    frameMostrado.grid(row = 0, column = 0)

    
def mostrarMayorCuenta2(datos, nombreCuenta,top, fr,totales,saldos):
    
    
    #nombreCuentaTextLabel = tk.Label(fr, text = nombreCuenta)
    #nombreCuentaTextLabel.grid(row = 1, column = 4)
    
    columns = ('#1', '#2', '#3','#4','#5','#6','#7')
    
    tree = ttk.Treeview(fr, height = 10, columns=columns, show='headings')
    
    tree.heading('#1', text='fecha')
    tree.heading('#2', text='numero comprobante')
    tree.heading('#3', text='tipo')
    tree.heading('#4', text='glosa')
    tree.heading('#5', text='debe')
    tree.heading('#6', text='haber')
    tree.heading('#7', text='saldo')
    
    tree.column("# 1", anchor = "center", width=100)
    tree.column("# 2", anchor = "center", width=130)
    tree.column("# 3", anchor = "center", width=100)
    tree.column("# 4", anchor = "center", width=130)
    tree.column("# 5", anchor = "center", width=100)
    tree.column("# 6", anchor = "center", width=100)
    tree.column("# 7", anchor = "center", width=100)
    
    tree.grid(row=0, columnspan= 30, sticky='nsew')
    
    totalSaldo = 0
    totalDebe = 0
    totalHaber = 0
    
    datosAux = []
    
    for item in datos:
        totalSaldo += (item[4] - item[5])
        datosAux.append((item[0], item[1], item[2], item[3], item[4], item[5], totalSaldo))
        totalDebe += item[4]   
        totalHaber += item[5]
        saldos.append(totalSaldo)
        
    for d in datosAux:
        tree.insert('', tk.END, values=d)
        
    scrollbar = ttk.Scrollbar(fr, orient=tk.VERTICAL, command=tree.yview)
    tree.configure(yscroll=scrollbar.set)
    scrollbar.grid(row=0, column=31, sticky='ns')
    
    totalNada = tk.Label(fr, text = "")
    totalNada.grid(row = 1, column = 4)
    
    totalDebeTextlabel = tk.Label(fr, text = "total debe: ")
    totalDebeTextlabel.grid(row = 2, column = 28)
    
    totalDebeLabel = tk.Label(fr, text = str(totalDebe))
    totalDebeLabel.grid(row = 2, column = 29)
    
    totalHaberTextlabel = tk.Label(fr, text = "total haber: ")
    totalHaberTextlabel.grid(row = 3, column = 28)
    
    totalHaberLabel = tk.Label(fr, text = str(totalHaber))
    totalHaberLabel.grid(row = 3, column = 29)
    
    totales.append(totalDebe)
    totales.append(totalHaber)
    
def proxiMayor(rut,fechaInicio,fechaFinal,datos,tipo,frames):
    if tipo.get() == "todos":
        VerMayor(rut,fechaInicio,fechaFinal,tipo.get(),frames)
    else:
        VerMayor(rut,fechaInicio,fechaFinal,datos[tipo.get().split()[0]][1],frames)

def MenuEmpresa(frames,datos):
    
    fr = frames["menuEmpresa"]
    titulo = tk.Label(fr, text = "menu de la empresa")
    titulo.grid(row = 0, column = 0)
    
    nombre, rut, duenyo, rutPropietario , direccion, giro = datos
    
    nombreLabel = tk.Label(fr, text = "nombre: " + nombre)
    nombreLabel.grid(row = 0, column = 0)
    
    rutLabel = tk.Label(fr, text = "rut: " + rut)
    rutLabel.grid(row = 0, column = 1)
    
    duenyoLabel = tk.Label(fr, text = "representante legal: " + duenyo)
    duenyoLabel.grid(row = 0, column = 2)
    
    rutduenyoLabel = tk.Label(fr, text = "rut del representante: " + duenyo)
    rutduenyoLabel.grid(row = 0, column = 3)
    
    direccionLabel = tk.Label(fr, text = ": " + duenyo)
    direccionLabel.grid(row = 0, column = 4)
    
    giroLabel = tk.Label(fr, text = "propietario: " + duenyo)
    giroLabel.grid(row = 0, column = 5)
    
    crearCuentaButton = tk.Button(fr,text = "crear cuenta", command = lambda: CrearCuenta(rut))
    crearCuentaButton.grid(row = 1, column = 0)
    
    editarCuentaButton = tk.Button(fr,text = "editar cuenta", command = lambda: SeleccionarCuenta(rut,"editar"))
    editarCuentaButton.grid(row = 2, column = 0)
    
    borrarCuentaButton = tk.Button(fr,text = "borrar cuenta", command = lambda: SeleccionarCuenta(rut,"borrar"))
    borrarCuentaButton.grid(row = 3, column = 0)
    
    verCuentaButton = tk.Button(fr,text = "ver cuentas", command = lambda: VerCuentas(rut))
    verCuentaButton.grid(row = 4, column = 0)
    
    crearComprobanteButton = tk.Button(fr,text = "crear comprobante", command = lambda: CrearComprobante(rut))
    crearComprobanteButton.grid(row = 5, column = 0)
    
    editarComprobanteButton = tk.Button(fr,text = "editar comprobante", command = lambda: SeleccionarComprobante(rut,"editar"))
    editarComprobanteButton.grid(row = 6, column = 0)
    
    borrarComprobanteButton = tk.Button(fr,text = "borrar comprobante", command = lambda: SeleccionarComprobante(rut,"borrar"))
    borrarComprobanteButton.grid(row = 7, column = 0)
    
    verComprobanteButton = tk.Button(fr,text = "ver comprobante", command = lambda: SeleccionarComprobante(rut,"ver"))
    verComprobanteButton.grid(row = 8, column = 0)
    
    diarioButton = tk.Button(fr,text = "ver diario", command = lambda: SeleccionarFecha(rut,"diario",frames))
    diarioButton.grid(row = 9, column = 0)
    
    mayorButton = tk.Button(fr,text = "ver mayor", command = lambda: SeleccionarFecha(rut,"mayor",frames))
    mayorButton.grid(row = 10, column = 0)
    
    balanceButton = tk.Button(fr,text = "ver balance", command = lambda: SeleccionarFecha(rut,"balance",frames))
    balanceButton.grid(row = 11, column = 0)
    
    frames["aux"].grid_forget()
    frames["aux"] = fr
    frames["aux"].grid(row = 0, column = 0)

def ventanaInicio(frames):
    fr = frames["inicio"]
    label = tk.Label(fr, text = "pantalla de inicio")
    label.grid(row = 0, column = 0)
    
    combo = ttk.Combobox(fr, state="readonly")
    combo.place(x=1, y=0)
    
    data = app.MostrarEmpresas()
    
    opciones = []
    
    dictEmpresas = {}
    
    for item in data:
        opciones.append(item[1] + " " + item[0])
        dictEmpresas[item[1]] = (item[0],item[1],item[2],item[3],item[4],item[5])
        
    combo["values"] = opciones

    ingresarButton = tk.Button(fr,text = "ingresar", command = lambda: MenuEmpresa(frames,dictEmpresas[combo.get().split()[0]]))
    ingresarButton.grid(row = 2,column = 2)

    editarButton = tk.Button(fr,text = "editar", command = lambda: EditarEmpresa(dictEmpresas[combo.get().split()[0]]))
    editarButton.grid(row = 2, column = 3)
    
    borrarButton = tk.Button(fr,text = "borrar", command = lambda: recagarBorrarEmpresa(data,combo,dictEmpresas,combo.get().split()[0]))
    borrarButton.grid(row = 2, column = 4)  
    
    crearButton = tk.Button(fr,text = "crear", command = lambda: crearEmpresa(data,combo,dictEmpresas))
    crearButton.grid(row = 2, column = 5)
    
    frames["aux"].grid_forget()
    frames["aux"] = fr
    frames["aux"].grid(row = 0, column = 1)
    
def CrearVentana():
    window = tk.Tk()
    window.title("contabilidad")
    
    return window

def CrearFrames(ventana,frames):
    inicio = tk.LabelFrame(ventana, text = "Inicio", padx = 5, pady = 5)
    #inicio.grid(row = 0, column = 1)
    frames["inicio"] = inicio
    
    frames["aux"] = inicio
    
    menuEmpresa = tk.LabelFrame(ventana, text = "menu de la empresa", padx = 5, pady = 5)
    frames["menuEmpresa"] = menuEmpresa
    
    ingresarComprobante = tk.LabelFrame(ventana, text = "menu del comprobante", padx = 5, pady = 5)
    frames["ingresarComprobante"] = ingresarComprobante
    
def obtenerCodigo(elemen):
    return elemen[3]
    
def VerBalance(rut,fechaInicio,fechaFinal):
    top = tk.Toplevel()
    top.title("Balance")
    datos2 = app.ObtenerBalance(rut,fechaInicio,fechaFinal)
    
    columns = ('#1', '#2', '#3','#4','#5','#6','#7','#8','#9','#10')
    
    tree = ttk.Treeview(top, height = 10, columns=columns, show='headings')
    
    tree.heading('#1', text='codigo cuenta')
    tree.heading('#2', text='nombre cuenta')
    tree.heading('#3', text='debitos')
    tree.heading('#4', text='creditos')
    tree.heading('#5', text='deudor')
    tree.heading('#6', text='acreedor')
    tree.heading('#7', text='activo')
    tree.heading('#8', text='pasivo')
    tree.heading('#9', text='perdida')
    tree.heading('#10', text='ganancia')
    
    tree.column("# 1", anchor = "center", width=100)
    tree.column("# 2", anchor = "center", width=140)
    tree.column("# 3", anchor = "center", width=100)
    tree.column("# 4", anchor = "center", width=100)
    tree.column("# 5", anchor = "center", width=100)
    tree.column("# 6", anchor = "center", width=100)
    tree.column("# 7", anchor = "center", width=100)
    tree.column("# 8", anchor = "center", width=100)
    tree.column("# 9", anchor = "center", width=100)
    tree.column("# 10", anchor = "center", width=100)
    
    tree.grid(row=0, columnspan= 30, sticky='nsew')
    
    datos = []
    
    for d in datos2:
        datos.append(d)
    
    datos.sort(key=obtenerCodigo)
    
    totalDebitos = 0
    totalCreditos = 0
    
    totalDeudor = 0
    totalAcreedor = 0
    
    totalActivo = 0
    totalPasivo = 0
    
    totalPerida = 0
    totalGanacia = 0
    
    datosExcel = {}
    
    datosCodigoCuenta = []
    datosNombreCuenta = []
    datosDebito = []
    datosCredito = []
    datosDeudor = []
    datosAcreedor = []
    datosActivo = []
    datosPasivo = []
    datosPerdida = []
    datosGanancia = []
    
    datosExcelTotales = {}
    
    datosAux = []
    
    for item in datos:
        
        deudor = item[0] - item[1]
        acreedor = item[1] - item[0]
        
        debitos = item[0]
        creditos = item[1]
        
        if deudor < 0:
            deudor = 0
            
        if acreedor < 0:
            acreedor = 0
            
        activo = 0
        pasivo = 0
        perdida = 0
        ganancia = 0
        
        if item[4] == "activo" or item[4] == "pasivo" or item[4] == "patrimonio":
            activo = deudor
            pasivo = acreedor
        else:
            perdida = deudor
            ganancia = acreedor
            
        totalCreditos += creditos
        totalDebitos += debitos
        
        totalDeudor += deudor
        totalAcreedor += acreedor
        
        totalActivo += activo
        totalPasivo += pasivo
        
        totalPerida += perdida
        totalGanacia += ganancia
        
        datosAux.append((item[3], item[2], item[0], item[1], deudor, acreedor, activo, pasivo, perdida, ganancia))
        datosCodigoCuenta.append(str(item[3]))
        datosNombreCuenta.append(str(item[2]))
        datosDebito.append(str(item[0]))
        datosCredito.append(str(item[1]))
        datosDeudor.append(str(deudor))
        datosAcreedor.append(str(acreedor))
        datosActivo.append(str(activo))
        datosPasivo.append(str(pasivo))
        datosPerdida.append(str(perdida))
        datosGanancia.append(str(ganancia))
        
    datosAux.append(("", "sub total:", str(totalDebitos), str(totalCreditos), str(totalDeudor), str(totalAcreedor), str(totalActivo), str(totalPasivo), str(totalPerida), str(totalGanacia)))
    datosAux.append(("", "resultado del ejercicio: ", "", "", "", "", str(abs(totalActivo - totalPasivo)), "", str(abs(totalPerida - totalGanacia)), ""))
        
    for d in datosAux:
        tree.insert('', tk.END, values=d)
        
    scrollbar = ttk.Scrollbar(top, orient=tk.VERTICAL, command=tree.yview)
    tree.configure(yscroll=scrollbar.set)
    scrollbar.grid(row=0, column=31, sticky='ns')
    
    datosExcel["codigoCuenta"] = datosCodigoCuenta
    datosExcel["nombreCuenta"] = datosNombreCuenta
    datosExcel["debito"] = datosDebito
    datosExcel["credito"] = datosCredito
    datosExcel["deudor"] = datosDeudor
    datosExcel["acreedor"] = datosAcreedor
    datosExcel["activo"] = datosActivo
    datosExcel["pasivo"] = datosPasivo
    datosExcel["perdida"] = datosPerdida
    datosExcel["ganancia"] = datosGanancia
    
    datosExcelTotales["totalDebitos"] = totalDebitos
    datosExcelTotales["totalCretidos"] = totalCreditos
    datosExcelTotales["totalDeudor"] = totalDeudor
    datosExcelTotales["totalAcreedor"] = totalAcreedor
    datosExcelTotales["totalActivo"] = totalActivo
    datosExcelTotales["totalPasivo"] = totalPasivo
    datosExcelTotales["totalPerida"] = totalPerida
    datosExcelTotales["totalGanacia"] = totalGanacia
    datosExcelTotales["totalDeudor"] = totalDeudor
    
    resultadoPerdidaGanancia = tk.Label(top,text = "")
    resultadoPerdidaGanancia.grid(row = 1, column = 0)
    
    #datos,datosTotales,razon,direccion,giro,representante,rut,rutRepresentante
    generar = tk.Button(top, text = "generar excel", command = lambda: GenerarExcelBalance(datosExcel,datosExcelTotales,rut,fechaInicio,fechaFinal))
    generar.grid(row = 2, column = 1,pady=5)

def GenerarExcelBalance(datos,datosTotales,rut,fechaInicio,fechaFinal):
    
    razon, direccion, giro, representante, rutRepresentante = app.ObtenerDatosParaExcel(rut)
    
    wb, hoja, thin_border, up_thin_border = PrepararExcel(rut,'balance')
    
    sheet = wb.get_sheet_by_name('balance')
    
    fechaInicioAux = fechaInicio.split('/')
    fechaFinalAux = fechaFinal.split('/')
    
    sheet['B2'].value = "BALANCE DE 8 COLUMNAS"
    sheet['B3'].value = "periodo: entre " + fechaInicioAux[1] + "/" + fechaInicioAux[0] + "/" + fechaInicioAux[2] + " y " + fechaFinalAux[1] + "/" + fechaFinalAux[0] + "/" + fechaFinalAux[2]
    sheet.merge_cells(start_row=2, start_column=2, end_row=2, end_column=9)
    
    sheet['B4'].value = "razon social: " + str(razon)
    sheet['B5'].value = "direccion: " + str(direccion)
    sheet['B6'].value = "giro: " + str(giro)
    sheet['B7'].value = "repre. legal: " + str(representante)
    sheet['F4'].value = "R.U.T.: " + str(rut)
    sheet['F5'].value = "comuna: " + str(giro)
    sheet['F6'].value = "ciudad: " + str(giro)
    sheet['F7'].value = "R.U.T.: " + str(rutRepresentante)
    sheet.merge_cells(start_row=4, start_column=2, end_row=4, end_column=3)
    sheet.merge_cells(start_row=5, start_column=2, end_row=5, end_column=3)
    sheet.merge_cells(start_row=6, start_column=2, end_row=6, end_column=3)
    sheet.merge_cells(start_row=7, start_column=2, end_row=7, end_column=3)
    sheet.merge_cells(start_row=4, start_column=6, end_row=4, end_column=7)
    sheet.merge_cells(start_row=5, start_column=6, end_row=5, end_column=7)
    sheet.merge_cells(start_row=6, start_column=6, end_row=6, end_column=7)
    sheet.merge_cells(start_row=7, start_column=6, end_row=7, end_column=7)
    
    
    sheet['B9'].value = "cuenta contables"
    sheet['C9'].value = "valores acumulados"
    sheet['E9'].value = "saldos"
    sheet['G9'].value = "inventario"
    sheet['I9'].value = "resultados"
    sheet.merge_cells(start_row=9, start_column=3, end_row=9, end_column=4)
    sheet.merge_cells(start_row=9, start_column=5, end_row=9, end_column=6)
    sheet.merge_cells(start_row=9, start_column=7, end_row=9, end_column=8)
    sheet.merge_cells(start_row=9, start_column=9, end_row=9, end_column=10)
    sheet.merge_cells(start_row=9, start_column=2, end_row=10, end_column=2)
    
    
    sheet['C10'].value = "debito"
    sheet['D10'].value = "credito"
    sheet['E10'].value = "deudor"
    sheet['F10'].value = "acreedor"
    sheet['G10'].value = "activo"
    sheet['H10'].value = "pasivo"
    sheet['I10'].value = "perdida"
    sheet['J10'].value = "ganancia"
    
    
    sheet['A9'].border = thin_border
    sheet['A10'].border = thin_border
    sheet['B9'].border = thin_border
    sheet['B10'].border = thin_border
    sheet['C9'].border = thin_border
    sheet['C10'].border = thin_border
    sheet['D9'].border = thin_border
    sheet['D10'].border = thin_border
    sheet['E9'].border = thin_border
    sheet['E10'].border = thin_border
    sheet['F9'].border = thin_border
    sheet['F10'].border = thin_border
    sheet['G9'].border = thin_border
    sheet['G10'].border = thin_border
    sheet['H9'].border = thin_border
    sheet['H10'].border = thin_border
    sheet['I9'].border = thin_border
    sheet['I10'].border = thin_border
    sheet['J9'].border = thin_border
    sheet['J10'].border = thin_border
    
    i = 11
    columnas = ['A','B','C','D','E','F','G','H','I','J']
    j = 0
    for key in datos:
        i = 11
        for d in datos[key]:
            sheet[columnas[j] + str(i)].value = d
            i += 1
            
        j += 1
        
  
    hoja.column_dimensions['A'].width = 13
    hoja.column_dimensions['B'].width = 15
    
    i += 3
    sheet['B' + str(i)].value = "subtotal"
    j = 2
    
    for key in datosTotales:
        sheet[columnas[j] + str(i)].value = int(datosTotales[key])
        j += 1
    
    sheet['B' + str(i)].border = up_thin_border
    sheet['C' + str(i)].border = up_thin_border
    sheet['D' + str(i)].border = up_thin_border
    sheet['E' + str(i)].border = up_thin_border
    sheet['F' + str(i)].border = up_thin_border
    sheet['G' + str(i)].border = up_thin_border
    sheet['H' + str(i)].border = up_thin_border
    sheet['I' + str(i)].border = up_thin_border
    sheet['J' + str(i)].border = up_thin_border
    
    i += 1
    
    inventario = sheet['G' + str((i - 1))].value - sheet['H' + str((i - 1))].value
    resultado = sheet['I' + str((i - 1))].value - sheet['J' + str((i - 1))].value
    
    sheet['B' + str(i)].value = "resultado del ejercicio"
    
    if inventario > 0:
        sheet['H' + str(i)].value = inventario
    else:
        sheet['G' + str(i)].value = inventario * (-1)
        
    if resultado > 0:
        sheet['J' + str(i)].value = resultado
    else:
        sheet['I' + str(i)].value = resultado * (-1)
        
    i += 1
    sheet['B' + str(i)].value = "totales"
    
    sheet['B' + str(i)].border = up_thin_border
    sheet['C' + str(i)].border = up_thin_border
    sheet['D' + str(i)].border = up_thin_border
    sheet['E' + str(i)].border = up_thin_border
    sheet['F' + str(i)].border = up_thin_border
    sheet['G' + str(i)].border = up_thin_border
    sheet['H' + str(i)].border = up_thin_border
    sheet['I' + str(i)].border = up_thin_border
    sheet['J' + str(i)].border = up_thin_border
    
    sheet['B' + str(i + 1)].border = up_thin_border
    sheet['C' + str(i + 1)].border = up_thin_border
    sheet['D' + str(i + 1)].border = up_thin_border
    sheet['E' + str(i + 1)].border = up_thin_border
    sheet['F' + str(i + 1)].border = up_thin_border
    sheet['G' + str(i + 1)].border = up_thin_border
    sheet['H' + str(i + 1)].border = up_thin_border
    sheet['I' + str(i + 1)].border = up_thin_border
    sheet['J' + str(i + 1)].border = up_thin_border
    
    j = 2
    for d in range(8):
        
        if sheet[columnas[j] + str(i - 1)].value != None:
            sheet[columnas[j] + str(i)].value = sheet[columnas[j] + str(i - 2)].value + sheet[columnas[j] + str(i - 1)].value
        
        else:
            sheet[columnas[j] + str(i)].value = sheet[columnas[j] + str(i - 2)].value
        
        #sheet[columnas[j] + str(i)].value = sheet[columnas[j] + str(i - 2)].value + sheet[columnas[j] + str(i - 1)].value
        
        j += 1
        
    i += 2
    sheet['C' + str(i)].value = "conforme a los dispuesto en el Art.100 inciso segundo del Codigo Trubutario, declaro que los asientos"
    sheet.merge_cells(start_row=i, start_column=3, end_row=i, end_column=12)
    i += 1
    sheet['C' + str(i)].value = "incorporados en el siguiente balance, corresponden a datos filedignos proporcionados por el"
    sheet.merge_cells(start_row=i, start_column=3, end_row=i, end_column=11)
    i += 1
    sheet['C' + str(i)].value = "contribuyente y/o representante legal de"
    sheet.merge_cells(start_row=i, start_column=3, end_row=i, end_column=6)
    sheet['G' + str(i)].value = str(razon)
    
    i += 4
    
    sheet['H' + str(i)].value = str(razon)
    sheet.merge_cells(start_row=i, start_column=8, end_row=i, end_column=10)
    
    i += 1
    sheet['C' + str(i)].value = "FIRMA CONTADOR"
    
    sheet['H' + str(i)].value = "FIRMA CONTRIBUYENTE"
    sheet.merge_cells(start_row=i, start_column=8, end_row=i, end_column=9)
    i += 1
    sheet['H' + str(i)].value = "R.U.T.  : " + str(rutRepresentante)
    sheet.merge_cells(start_row=i, start_column=8, end_row=i, end_column=9)
    
    nombre = str(razon) + " " + str(fechaInicio) + " " + str(fechaFinal) + "balance.xlsx"
    nombre = nombre.replace("/", "-")
    
    wb.save(nombre)
    
    print("excel creado")

def GenerarExcelMayor(datos,datosTotales,saldos,rut,fechaInicio,fechaFinal):
    wb, hoja, thin_border, up_thin_border = PrepararExcel(rut,'mayor')
    sheet = wb.get_sheet_by_name('mayor')
    
    fechaInicioAux = fechaInicio.split('/')
    fechaFinalAux = fechaFinal.split('/')
    
    i = 2
    
    
    for c in datos:
        auxList = datos[c]
        sald = saldos[c]
        
    
        sheet['B' + str(i)].value = "MAYOR GENERAL DE " + str(auxList[0][7].upper())
        i += 1
        sheet['B' + str(i)].value = "periodo: entre " + fechaInicioAux[1] + "/" + fechaInicioAux[0] + "/" + fechaInicioAux[2] + " y " + fechaFinalAux[1] + "/" + fechaFinalAux[0] + "/" + fechaFinalAux[2]
        sheet.merge_cells(start_row=2, start_column=2, end_row=2, end_column=9)
        
        i += 2
        
        sheet['B' + str(i)].value = "fecha"
        sheet['C'  + str(i)].value = "comprobante"
        sheet['D'  + str(i)].value = "tipo"
        sheet['E'  + str(i)].value = "glosa"
        sheet['F'  + str(i)].value = "debe"
        sheet['G'  + str(i)].value = "haber"
        sheet['H'  + str(i)].value = "saldo"
        
        i += 1
        
        sheet['B' + str(i)].border = up_thin_border
        sheet['C' + str(i)].border = up_thin_border
        sheet['D' + str(i)].border = up_thin_border
        sheet['E' + str(i)].border = up_thin_border
        sheet['F' + str(i)].border = up_thin_border
        sheet['G' + str(i)].border = up_thin_border
        sheet['H' + str(i)].border = up_thin_border
    
        columnas = ['B','C','D','E','F','G','H']
        i += 1
        index = 0
        for item in auxList:
            j = 0
            auxItem = (item[0],item[1],item[2],item[3],item[4],item[5],sald[index])
            index += 1
            for a in auxItem:
                sheet[columnas[j] + str(i)].value = str(a)
                j += 1
                
            i += 1
            
        sheet['E' + str(i)].value = "TOTAL"
    
        j = 4
        
        totales = datosTotales[c]
        
        for item in totales:
            sheet[columnas[j] + str(i)].value = str(item)
            
            j += 1
            
        for c in columnas:
            sheet[c + str(i)].border = up_thin_border
            
    
        i += 3
        
    
    nombre = str(rut) + " " + str(fechaInicio) + " " + str(fechaFinal) + " mayor.xlsx"
    nombre = nombre.replace("/", "-")
    
    wb.save(nombre)
    
    print("excel creado")



def main():

    window = CrearVentana()
    frames = {}
    
    CrearFrames(window, frames)
    #btn = tk.Button(window, text = "ingresar", command = lambda: texto(window))
    #btn.grid(row = 1, column = 0)
    ventanaInicio(frames)
    window.mainloop()
    
main()